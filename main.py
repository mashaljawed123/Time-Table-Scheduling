#Logic 
# This code is designed to create a weekly timetable for courses.
# It works by first defining various sections and courses offered for different years 
# (1st year, 3rd year, 5th year, 7th year),
# then assigns rooms and timeslots to each section and course.
# The sections and courses are represented as a bipartite graph, 
# where one side contains sections and the other side contains courses. 
# The code uses maximum matching to allocate each course to a section, 
# ensuring that no section is assigned the same course at the same time. 
# Then, it generates a timetable for each weekday and saves it in an Excel workbook, with different colors 
# for different semesters.
# Import necessary libraries
import networkx as nx  # Import the NetworkX library for graph operations
from itertools import product  # Import the product function from itertools to create combinations
from openpyxl import Workbook  # Import Workbook from openpyxl to create Excel files
from openpyxl.styles import PatternFill  # Import PatternFill to add color in Excel cells
import matplotlib.pyplot as plt  # Import matplotlib for plotting graphs

# Define data
se_sections_7 = ['7AS', '7BS', '7CS', '7DS', '7ES']  # List of sections for semester 7
se_courses_7 = ['FY-1', 'PPIT', 'SE', 'RS', 'MM']  # List of courses for semester 7
se_comb_7 = list(product(se_sections_7, se_courses_7))  # Create all combinations of sections and courses for semester 7

sections_7 = ['7A', '7B', '7C', '7D', '7E']  # List of sections for semester 7
courses_7 = ['FYP-1', 'IR', 'PSYC', 'DLP', 'FSPM']  # List of courses for semester 7
comb_7 = list(product(sections_7, courses_7))  # Create all combinations of sections and courses for semester 7

sections_5 = ['5A', '5B', '5C', '5D', '5E']  # List of sections for semester 5
courses_5 = ['PDC', 'DB', 'ALGO', 'GT', 'SDA']  # List of courses for semester 5
comb_5 = list(product(sections_5, courses_5))  # Create all combinations of sections and courses for semester 5

sections_3 = ['3A', '3B', '3C', '3D', '3E']  # List of sections for semester 3
courses_3 = ['COAL', 'DSA', 'DS', 'LA', 'POE']  # List of courses for semester 3
comb_3 = list(product(sections_3, courses_3))  # Create all combinations of sections and courses for semester 3

sections_1 = ['1A', '1B', '1C', '1D', '1E']  # List of sections for semester 1
courses_1 = ['PF', 'ICT', 'IRS', 'CAL', 'PS', 'AP']  # List of courses for semester 1
comb_1 = list(product(sections_1, courses_1))  # Create all combinations of sections and courses for semester 1

sections_courses = comb_1 + comb_3 + comb_5 + comb_7 + se_comb_7  # Combine all section-course combinations into one list
rooms = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'R11','R12','R109']  # List of available rooms
timeslots = ['8-9', '9-10', '10-11', '11-12', '12-1', '1-2', '2-3', '3-4']  # List of time slots for each day
rooms_slots = list(product(rooms, timeslots))  # Create all combinations of rooms and timeslots
week_days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']  # List of weekdays

semester_colors = {  # Dictionary to store colors for each semester
    '7': "FFFF00",  # Yellow for semester 7
    '5': "FF00FF",  # Magenta for semester 5
    '3': "FFCCCB",  # Light pink for semester 3
    '1': "CCE5FF"   # Light blue for semester 1
}

# 1. Section-to-Course Bipartite Graph
G = nx.Graph()  # Create an empty graph G using NetworkX
for sec, course in sections_courses:  # Loop through all section-course combinations
    G.add_node(sec, bipartite=0)  # Add section as a node in the graph
    G.add_node(course, bipartite=1)  # Add course as a node in the graph
    G.add_edge(sec, course)  # Create an edge between the section and the course

# Adjust positions for Section-Course graph
pos = {
    node: (0, i * 15) for i, node in enumerate
    (sections_1 + sections_3 + sections_5 + se_sections_7 + sections_7)
    }  # Set positions for sections
pos.update({
    node: (1, i * 15) for i, node in enumerate
    (courses_1 + courses_3 + courses_5 + se_courses_7 + courses_7)
    })  # Set positions for courses

plt.figure(figsize=(12, 18))  # Set the figure size for the plot
nx.draw(G, pos, with_labels=True, font_weight="bold", node_color="lightgreen", node_size=1500)  # Draw the bipartite graph
plt.title("Bipartite Graph of Sections and Courses")  # Set the title of the graph
plt.show()  # Display the graph

# 2. Create workbook for daily timetables
wb = Workbook()  # Create a new Excel workbook

# Helper function to draw bipartite graphs
def draw_bipartite_graph(G, pos, matching_edges, title):
    plt.figure(figsize=(12, 8))  # Set figure size for the plot
    nx.draw(  # Draw the bipartite graph
        G,
        pos,
        with_labels=True,
        font_weight="bold",
        node_color="lightgreen",
        node_size=1500,
        edge_color="black"
    )
    # Highlight matching edges in red
    nx.draw_networkx_edges(G, pos, edgelist=matching_edges, edge_color="red", width=2)
    plt.title(title)  # Set the title of the graph
    plt.show()  # Display the graph

# Process each day
for day in week_days:  # Loop through each day of the week
    ws = wb.create_sheet(title=day)  # Create a new sheet in the Excel workbook for each day
    ws.append(["Timeslots"] + rooms)  # Add the header row with timeslots and rooms

    # Graph for Section-Course
    G_section_course = nx.Graph()  # Create a new bipartite graph for section-course combinations
    G_section_course.add_nodes_from(sections_courses, bipartite=1)  # Add section-course nodes to the graph
    G_section_course.add_nodes_from(rooms_slots, bipartite=0)  # Add room-time slot nodes to the graph
    G_section_course.add_edges_from(product(rooms_slots, sections_courses))  # Create edges between room-time slots and sections

    # Graph for Room-Timeslot
    G_room_timeslot = nx.Graph()  # Create a new bipartite graph for room-timeslot combinations
    G_room_timeslot.add_nodes_from(rooms, bipartite=0)  # Add room nodes to the graph
    G_room_timeslot.add_nodes_from(timeslots, bipartite=1)  # Add timeslot nodes to the graph
    G_room_timeslot.add_edges_from(product(rooms, timeslots))  # Create edges between rooms and timeslots

    # Maximum matching
    max_matching = nx.bipartite.maximum_matching(G_section_course)  # Find the maximum matching between sections and rooms
    valid_matching = {}  # Dictionary to store valid matches
    section_timeslot_allocated = {timeslot: [] for timeslot in timeslots}  # Track sections already allocated to each timeslot
    course_assigned = {section: set() for section, _ in sections_courses}  # Track assigned courses for each section

    for room_timeslot, course_section in max_matching.items():  # Loop through the matching results
        if isinstance(room_timeslot, tuple) and room_timeslot in rooms_slots:  # Check if the room-timeslot is valid
            room, timeslot = room_timeslot  # Unpack room and timeslot
            section, course = course_section  # Unpack section and course
            
            # Check if the section is already allocated to the timeslot and if the course has already been assigned to the section
            if section not in section_timeslot_allocated[timeslot] and course not in course_assigned[section]:
                valid_matching[room_timeslot] = course_section  # Add to valid matching
                section_timeslot_allocated[timeslot].append(section)  # Mark the section as allocated to the timeslot
                course_assigned[section].add(course)  # Mark the course as assigned to the section
            else:
                # If the section already has this course or the course is already assigned for the section, skip
                for new_section, new_course in sections_courses:  # Try to find a new section-course combination
                    if new_section not in section_timeslot_allocated[timeslot] and new_course not in course_assigned[new_section]:
                        valid_matching[room_timeslot] = (new_section, new_course)  # Assign a different section and course
                        section_timeslot_allocated[timeslot].append(new_section)# Mark the new section as allocated
                        course_assigned[new_section].add(new_course)# Mark the new course as assigned
                        break  # Stop once a valid section and course are found

    # Draw Section-Course Bipartite Graph for the Day
    pos_section_course = {
        node: (0, i * 90) for i, node in enumerate(rooms_slots)
    }
    pos_section_course.update({
        node: (1, i * 90) for i, node in enumerate(sections_courses)
    })

    plt.figure(figsize=(60, 50))
    nx.draw(
        G_section_course,
        pos_section_course,
        with_labels=True,
        font_weight="bold",
        node_color="lightgreen",
        node_size=3500,
        edge_color="black"
    )
    # Highlight matching edges in red
    nx.draw_networkx_edges(
        G_section_course,
        pos_section_course,
        edgelist=[(k, v) for k, v in valid_matching.items()],
        edge_color="red",
        width=2
    )
    plt.title(f"Bipartite Graph of Sections and Courses for {day}")
    plt.show()

    # Draw Room-Timeslot Bipartite Graph
    pos_room_timeslot = {
        node: (0, i * 15) for i, node in enumerate(rooms)
    }
    pos_room_timeslot.update({
        node: (1, i * 15) for i, node in enumerate(timeslots)
    })
    draw_bipartite_graph(
        G_room_timeslot, pos_room_timeslot, [], f"{day} Room-Timeslot Bipartite Graph"
    )

    # Populate Excel sheet
    timetable_data = {timeslot: [""] * len(rooms) for timeslot in timeslots}
    for room_timeslot, course_section in valid_matching.items():
        room, timeslot = room_timeslot
        section, course = course_section
        room_index = rooms.index(room)
        timetable_data[timeslot][room_index] = f"{course} {section}"

    for i, timeslot in enumerate(timeslots):
        row = [timeslot] + timetable_data[timeslot]
        ws.append(row)
        for j, cell_value in enumerate(row[1:], start=2):
            if cell_value:
                semester = cell_value.split(" ")[1][0]
                color = semester_colors.get(semester, "FFFFFF")
                ws.cell(row=i + 2, column=j).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

# Save workbook
wb.remove(wb["Sheet"])
wb.save("Weekly_Timetable.xlsx")
